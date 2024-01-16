from openpyxl import Workbook, load_workbook 
wb=load_workbook('colors.xlsx')
ws=wb['Sheet1']
ws1 = wb['Done']
max_row=ws.max_row

colors = []
n = int(input("Cik daudz krāsas gribi ievadīt?: "))
 
for i in range(0, n):
    number = int(input("Ievadi krāsu: "))
    colors.append(number)



row_num = 1
for each in colors:
    done = ws1.cell(row=row_num, column=1, value=each)
    row_num += 1
    done1 = done.value

wb.save('colors.xlsx')

# col_kopa = []
data = []
kopa = 0
for row in range(2, max_row+1):
    num = ws['A'+str(row)].value
    stitches = ws['D'+str(row)].value

    # col = ws['B' +str(row)].value
    # if type(col) == str and type(stitches) == int:                # ar rindiņām, kuras ir iekomentētas, var noteikt cik procenti būs paveikti, 
    #     if col == "red":                                          #           ja izdarīs visas izvēlētās, piemēram, sarkanas krāsas.
    #         col_kopa.append(stitches)
            


    
    if type(stitches) == int:
        kopa += stitches
        if num in colors:
            data.append(stitches)


# col_sum = sum(col_kopa)
# procenti = (col_sum/kopa)*100

summa = sum(data)
procenti = (summa/kopa)*100
print("Tu esi izdaríjis/usi",round(procenti, 1), "% no darba")



wb.close()
